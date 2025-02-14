import os
from azure.core.credentials import AzureKeyCredential
from azure.ai.documentintelligence import DocumentIntelligenceClient
from azure.ai.documentintelligence.models import DocumentAnalysisFeature
from azure.core.exceptions import HttpResponseError
from dotenv import load_dotenv
import pandas as pd
import io
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.hyperlink import Hyperlink
from config import settings
from model import extract_invoice_details

config = settings.get_settings()
endpoint = config.document_intelligence.endpoint
key = config.document_intelligence.api_key

if not endpoint or not key:
    raise ValueError("DOCUMENTINTELLIGENCE_ENDPOINT and DOCUMENTINTELLIGENCE_API_KEY must be set in .env.")

client = DocumentIntelligenceClient(endpoint=endpoint, credential=AzureKeyCredential(key))

def extract_fields_from_invoice(file_path):
    """
    Extract specified fields from an invoice using Azure Document Intelligence.

    Args:
        file_path (str): Path to the invoice file.

    Returns:
        dict: Extracted fields with their values and confidence levels.
    """
    try:
        fields_to_extract = {
            "Vendor Name": "VendorAddressRecipient",
            # "Customer Name": "CustomerAddressRecipient",
            "Date": "InvoiceDate",
            # "Project Name": "ProjectName", 
            # "Job Name": "ShipToAddressRecipient",
            "Total Value": "InvoiceTotal",
            "Invoice Number": "InvoiceId"
        }
        with open(file_path, "rb") as document:
            poller = client.begin_analyze_document(
                model_id="prebuilt-invoice",
                body=document,
                features=[DocumentAnalysisFeature.QUERY_FIELDS],
                query_fields=[*list(fields_to_extract.values())],
            )
            result = poller.result()
        
        
        extracted_data = {}

        if result.documents:
            for document in result.documents:
                for field_name, model_field in fields_to_extract.items():
                    field = document.fields.get(model_field)
                    if field:
                        if model_field == "InvoiceTotal":
                            if field.value_currency:
                                extracted_data[field_name] = {
                                    "value": str(field.value_currency.amount) + " " + field.value_currency.currency_code,
                                    "confidence": field.confidence
                                }
                            else:
                                extracted_data[field_name] = {
                                    "value": field.value_string,
                                    "confidence": field.confidence
                                }
                        elif model_field == "InvoiceDate":
                            extracted_data[field_name] = {
                                "value": field.value_date,
                                "confidence": field.confidence
                            }
                        else:
                            extracted_data[field_name] = {
                                "value": field.content,
                                "confidence": field.confidence
                            }
        invoice_text=result.content

        return extracted_data,invoice_text

    except HttpResponseError as e:
        print(f"An error occurred: {e.message}")
        return None
    
def generate_invoice_excel(extracted_fields, invoice_text, invoice_path, save_dir="temp_uploads"):
    """
    Generates an Excel file from extracted invoice data and saves it to disk.

    Args:
        extracted_fields (dict): Extracted invoice fields with values.
        invoice_path (str): Path to the original invoice file.
        save_dir (str): Directory where the Excel file will be saved.

    Returns:
        str: Path to the saved Excel file.
    """
    
    extra_fields=extract_invoice_details(invoice_text, extracted_fields)

    # Ensure the save directory exists
    os.makedirs(save_dir, exist_ok=True)

    # Prepare data for the DataFrame
    data = {key: [details["value"]] for key, details in extracted_fields.items()}
    for key, value in extra_fields.items():
        data[key] = [value]
    data["Invoice Path"] = [f'=HYPERLINK("{invoice_path}", "Open Invoice")']

    # Create DataFrame
    df = pd.DataFrame(data)

    # Define file path
    excel_filename = "extracted_invoice.xlsx"
    excel_path = os.path.join(save_dir, excel_filename)

    # Save to an Excel file
    with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Invoice Data")

        # Access the workbook and sheet
        workbook = writer.book
        sheet = writer.sheets["Invoice Data"]

        # Define header style (green background, white text)
        header_fill = PatternFill(start_color="228B22", end_color="228B22", fill_type="solid")  # Green background
        header_font = Font(color="FFFFFF", bold=True)  # White text
        link_font = Font(color="0000FF", underline="single")  # Blue, underlined hyperlink text

        # Apply header styles and adjust column widths
        for col_idx, col in enumerate(df.columns, start=1):
            cell = sheet.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            sheet.column_dimensions[get_column_letter(col_idx)].width = 25  # Adjust column width

        # Apply hyperlink style to "Invoice Path" column
        invoice_path_col = list(df.columns).index("Invoice Path") + 1  # Get column index (1-based)
        sheet.cell(row=2, column=invoice_path_col).font = link_font  # Apply hyperlink style

    return excel_path  # Return the path of the saved file


if __name__ == "__main__":
    invoice_file_path = r"D:\work\upwork\armstrong\invoices\INVOICE # 43353_240826_131415.pdf"
    # invoice_file_path = r"D:\work\upwork\armstrong\2168515 ACB110.pdf"

    if not os.path.exists(invoice_file_path):
        raise FileNotFoundError(f"File not found: {invoice_file_path}")

    print(f"Analyzing invoice: {invoice_file_path}")

    extracted_fields, invoice_text = extract_fields_from_invoice(invoice_file_path)

    print(extract_invoice_details(invoice_text, extracted_fields))

    if extracted_fields:
        print("\nExtracted Fields:")
        for field, details in extracted_fields.items():
            print(f"{field}: {details['value']} (Confidence: {details['confidence']:.2f})")
    else:
        print("No fields extracted.")
