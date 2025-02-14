from concurrent.futures import ThreadPoolExecutor, as_completed
from cogservice import extract_fields_from_invoice
import os
import pandas as pd
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
from config import settings
from model import extract_invoice_details


config = settings.get_settings()
endpoint = config.document_intelligence.endpoint
key = config.document_intelligence.api_key


def batch_extract_fields_from_invoices(invoice_file_paths, parallel=True):
    """
    Processes a list of invoice files in batch and extracts their fields.

    Args:
        invoice_file_paths (list): List of file paths to the invoice files.
        parallel (bool): If True, process files concurrently.

    Returns:
        dict: A dictionary mapping each invoice file path to a tuple:
              (extracted_fields, invoice_text). If a file fails processing, its value will be None.
    """
    results = {}

    if parallel:
        # Process invoices concurrently using a ThreadPoolExecutor
        with ThreadPoolExecutor() as executor:
            future_to_file = {
                executor.submit(extract_fields_from_invoice, file_path): file_path
                for file_path in invoice_file_paths
            }
            for future in as_completed(future_to_file):
                file_path = future_to_file[future]
                try:
                    result = future.result()
                    results[file_path] = result
                except Exception as e:
                    print(f"Error processing {file_path}: {e}")
                    results[file_path] = None
    else:
        # Process invoices sequentially
        for file_path in invoice_file_paths:
            results[file_path] = extract_fields_from_invoice(file_path)

    return results
def generate_batch_invoices_excel(batch_results, save_dir="temp_uploads"):
    """
    Generates a single Excel workbook that summarizes the extracted data for multiple invoices.

    Each row in the Excel sheet corresponds to one invoice.

    Args:
        batch_results (dict): A dictionary where each key is an invoice file path and each value is
                              a tuple (extracted_fields, invoice_text).
        save_dir (str): Directory where the Excel file will be saved.

    Returns:
        str: Path to the saved Excel workbook.
    """
    os.makedirs(save_dir, exist_ok=True)
    rows = []

    for invoice_path, result in batch_results.items():
        if result is None:
            # Optionally, add a row indicating that processing failed for this invoice.
            rows.append({
                "Invoice File": invoice_path,
                "Error": "Processing failed"
            })
            continue

        extracted_fields, invoice_text = result
        # Get additional details using your custom extraction method.
        extra_fields = extract_invoice_details(invoice_text, extracted_fields)

        # Build a row with both the extracted and extra fields.
        row = {}
        for key, details in extracted_fields.items():
            row[key] = details["value"]
        # Merge any extra fields.
        row.update(extra_fields)
        # Add a hyperlink to the original invoice.
        row["Invoice Path"] = f'=HYPERLINK("{invoice_path}", "Open Invoice")'
        rows.append(row)

    # Create a DataFrame from the aggregated rows.
    df = pd.DataFrame(rows)

    # Define the output Excel file path.
    excel_filename = "batch_extracted_invoices.xlsx"
    excel_path = os.path.join(save_dir, excel_filename)

    # Write the DataFrame to Excel with header and hyperlink styling.
    with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Invoice Data")
        workbook = writer.book
        sheet = writer.sheets["Invoice Data"]

        header_fill = PatternFill(start_color="228B22", end_color="228B22", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        link_font = Font(color="0000FF", underline="single")

        # Apply header styles and adjust column widths.
        for col_idx, col in enumerate(df.columns, start=1):
            cell = sheet.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            sheet.column_dimensions[get_column_letter(col_idx)].width = 25

            # If the column is "Invoice Path", style all cells in that column as hyperlinks.
            if col == "Invoice Path":
                for row_idx in range(2, len(df) + 2):
                    sheet.cell(row=row_idx, column=col_idx).font = link_font

    return excel_path


# Sample usage of the batch methods.
if __name__ == "__main__":
    # Example: list of invoice file paths (update these paths as needed)
    invoice_files = [
        r"D:\work\upwork\armstrong\invoices\INVOICE # 43353_240826_131415.pdf",
        r"D:\work\upwork\armstrong\invoices\20231109144311839.pdf",
        # Add more file paths as needed.
    ]

    # Ensure that the files exist.
    invoice_files = [path for path in invoice_files if os.path.exists(path)]
    if not invoice_files:
        raise FileNotFoundError("None of the specified invoice files were found.")

    print("Starting batch extraction of invoice fields...")
    batch_results = batch_extract_fields_from_invoices(invoice_files, parallel=True)

    # Print a summary of extracted fields for each invoice.
    for file_path, result in batch_results.items():
        if result:
            extracted_fields, _ = result
            print(f"\nInvoice: {file_path}")
            for field, details in extracted_fields.items():
                print(f"  {field}: {details['value']} (Confidence: {details['confidence']:.2f})")
        else:
            print(f"\nInvoice: {file_path} - Extraction failed.")

    # Generate a consolidated Excel file for all processed invoices.
    excel_output_path = generate_batch_invoices_excel(batch_results, save_dir="temp_uploads")
    print(f"\nBatch invoice Excel generated at: {excel_output_path}")
